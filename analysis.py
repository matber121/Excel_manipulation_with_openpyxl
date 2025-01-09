import openpyxl
from openpyxl.chart import Reference,BarChart
from openpyxl.chart.layout import Layout, ManualLayout


wb = openpyxl.load_workbook(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

ws = wb['vgsales']

print('Total number of rows: ' + str(ws.max_row) + '. And total number of columns: ' + str(ws.max_column))

#Print value of  specific cell
print('print the value:'+ws['A1'].value)

#Reading Data From Multiple Cells

values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)

#Printing out multiple rows in a specific column
data= [ws.cell(row=i,column=2).value for i in range(2,12)]
print(data)

# reading data from a range of cells (from column 1 to 6)

my_list=list()

for i in   ws.iter_rows(
    min_row=1,max_row=11,min_col=1,max_col=6,
    values_only=True):
    my_list.append(i)

for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
    print ("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1,ele2,ele3,ele4,ele5,ele6))
     
# Writing to a Cell
ws['k1'] = 'sum of sales'
#wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

#Creating a New Column
row_position = 2
col_position = 7

total_sales = ((ws.cell(row=row_position, column=col_position).value)+
               (ws.cell(row=row_position, column=col_position+1).value)+
               (ws.cell(row=row_position, column=col_position+2).value)+
               (ws.cell(row=row_position, column=col_position+3).value))

ws.cell(row=2,column=11).value=total_sales
# wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

#  create a for loop to sum the sales values in every row

row_position = 1
for i in range(1,ws.max_row):
    row_position += 1
    NA_SALES = ws.cell(row=row_position,column=7).value
    EU_SALES = ws.cell(row=row_position,column=8).value
    JP_SALES = ws.cell(row=row_position,column=9).value
    OTHER_SALES = ws.cell(row=row_position,column=10).value

    total_sales = (NA_SALES+EU_SALES+JP_SALES+OTHER_SALES)
    ws.cell(row=row_position,column=7).value = total_sales

 

wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

# Appending New Rows
new_row = (1,'The Legend of Zelda',1986,'Action','Nintendo',3.74,0.93,1.69,0.14,6.51,6.5)

ws.append(new_row)
#wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

# printing the last row in the workbook 
values = [ws.cell(row=ws.max_row,column=i).value for i in range(1,ws.max_column+1)]
print(values)

# delete the last row
ws.delete_rows(ws.max_row, 1) # row number, number of rows to delete

# wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

# Creating Excel Formulas with Openpyxl
ws['P1'] = 'Average Sales'
ws['P2'] = '= AVERAGE(K2:K16220)'

ws['Q1'] = "Number of Populated Cells" 
ws['Q2'] = '=COUNTA(E2:E16220)'

ws['R1'] = 'Number of Rows with Sports Genre'
ws['R2'] = '=COUNTIF(E2:E16220, "Sports")'

ws['S1'] = 'Total Sports Sales'
ws['S2'] = '=SUMIF(E2:E16220, "Sports",K2:K16220)'
 
ws['T1'] = 'Rounded Sum of Sports Sales'
ws['T2'] = '=CEILING(S2,25)'

# wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')

# Adding Charts to an Excel File with Openpyxl
ws = wb['Total Sales by Genre']

values = Reference(ws,
                   min_col=2,
                   max_col=2,
                   min_row=2,
                   max_row=13)
cats = Reference(ws, 
                 min_col=1, 
                 max_col=1, 
                 min_row=2, 
                 max_row=13)


chart = BarChart()
chart.add_data(values,titles_from_data=True)
chart.set_categories(cats)

chart.title = 'Total Sales'
chart.x_axis.title = 'Genre'

chart.y_axis.title = 'Total sales by genre'

ws.add_chart(chart,"D2")

wb.save(r'C:\Users\alima\Desktop\Excel_manulpation\videogamesales.xlsx')


# Create the chart
chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)

# Set chart title and move it above the chart
chart.title = "Sales Breakdown"
chart.layout = Layout(
    ManualLayout(
        h=0.2,  # Adjust height of the chart title area
        x=0.25,  # Center the title horizontally
        y=0.05  # Position the title at the top
    )
)

# Set x-axis title and move it below the x-axis
chart.x_axis.title = "Genre"
chart.x_axis.title_layout = Layout(
    ManualLayout(
        x=0.5,  # Center the x-axis title
        y=1.0   # Position the x-axis title below the axis
    )
)

# Set y-axis title and move it to the side
chart.y_axis.title = "Breakdown of Sales by Genre"
chart.y_axis.title_layout = Layout(
    ManualLayout(
        x=-0.1,  # Position the y-axis title to the left of the chart
        y=0.5    # Center the y-axis title vertically
    )
)

# Add the chart to the worksheet
ws.add_chart(chart, "H2")

# Save the workbook
wb.save("videogamesales.xlsx")
